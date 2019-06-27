#=====================================================================================================
# ScritName          :
# Version            :
# Description        :
# Developed By       : Devs
# Developed On       :
# Last Modified      :
# Last Modified Date :
# Modified by        :
#======================================================================================================
import csv
import openpyxl
import time
import sys
import os

#user parameter
user=sys.argv[1]

start_time=time.time() #time before the start of the conversion

excel_file="/home/"+ user +"/config.xlsx"
text_file="/home/"+ user +"/config.txt"
testing_log ="/home/" + user + "/testing_log.log"
prev_row_count="/home/"+ user +"/prev_row_count.txt"

#Getting the row count of the excel file before any updation/modification
with open(prev_row_count, 'r') as file :
  txt_file = file.read()
lst=txt_file.split("=")
prev_row_cnt=int(lst[1])

#log file 
testing_log=open(testing_log,"a+")

#Opening the excel file for reading
wb = openpyxl.load_workbook(excel_file)
sheet = wb['Sheet1']

#Getting the latest row count of excel file
curr_row_cnt=sheet.max_row
#Row count of the newly added rows in excel
new_rows_cnt = curr_row_cnt - prev_row_cnt

if prev_row_cnt != curr_row_cnt :
	#converting the excel file to text file
	if new_rows_cnt == curr_row_cnt:
		if os.path.isfile(text_file):
			os.remove(text_file)
		with open(text_file , "a+") as file:
			writer = csv.writer(file, delimiter = "|")
			print sheet, sheet.max_row, sheet.max_column
			for row_idx in range(2, int(sheet.max_row)+1):
				row=[]
				for col_idx in range(1, int(sheet.max_column)+1):
					row.append(sheet.cell(row=row_idx, column=col_idx).value) 
				writer.writerow(row)
	#if the excel file gets updated , updates the text file with newly added rows
	else:
		with open(text_file , "a+") as file:
			writer = csv.writer(file, delimiter = "|")
			print sheet, sheet.max_row, sheet.max_column
			for row_idx in range((new_rows_cnt+prev_row_cnt+1), int(sheet.max_row)+1):
				row=[]
				for col_idx in range(1, int(sheet.max_column)+1):
					row.append(sheet.cell(row=row_idx, column=col_idx).value) 
				writer.writerow(row)

# Updating the prev_row_count.txt file with the current row count 
txt_file = txt_file.replace(str(prev_row_cnt), str(curr_row_cnt))
# Write the file out again
with open(prev_row_count, 'w') as file:
  file.write(txt_file)
end_time=time.time()
testing_log.write("Took %s seconds to convert excel file to text file" % (end_time - start_time))


		
		
		
		
		